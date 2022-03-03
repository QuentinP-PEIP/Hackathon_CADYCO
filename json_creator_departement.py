import openpyxl
import json

wookbook = openpyxl.load_workbook("Données/PoleEmploi/Extraction_OFF_1_Bassins_d_emploi.xlsx")

wookbook.active = 0

sheet  = wookbook.active

first_rows = []

for row in sheet.iter_rows(max_row=3):
    first_rows.append(row)

dict_number = {'38': {}, '69': {}, '73': {}, '74': {}}
dict_percent = {'38': {}, '69': {}, '73': {}, '74': {}}

row_data = []

for row in sheet.iter_rows(max_row=3):
    this_data = []
    for i in range(7, len(row)):       
        this_data.append(row[i].value)
    row_data.append(this_data)

row_data[2] = row_data[2][2:]



for row in sheet.iter_rows(min_row=4):
    if not dict_number.get(row[0].value).get(row[2].value):
        dict_number[row[0].value][row[2].value] = {'PE': {}, 'PE-CUMUL12MOIS': {}, 'TOFF': {}, 'TOFF-CUMUL12MOIS': {}}


index = 0
for row in sheet.iter_rows(min_row=4):
    for i in range(0, len(row_data[2]), 2):
           
        dict_number[row[0].value][row[2].value][row[6].value][row_data[2][i]] = row[i+9].value
    dict_number[row[0].value][row[2].value][row[6].value]["nbr_offres"] = row[7].value


for row in sheet.iter_rows(min_row=4):
    if not dict_percent.get(row[0].value).get(row[2].value):
        dict_percent[row[0].value][row[2].value] = {'PE': {}, 'PE-CUMUL12MOIS': {}, 'TOFF': {}, 'TOFF-CUMUL12MOIS': {}}


index = 0
for row in sheet.iter_rows(min_row=4):
    for i in range(0, len(row_data[2]), 2):
        dict_percent[row[0].value][row[2].value][row[6].value][row_data[2][i]] = row[i+10].value
    dict_percent[row[0].value][row[2].value][row[6].value]["nbr_offres"] = row[7].value

with open('Bassin_number.json', 'w') as outfile:
    json.dump(dict_number, outfile)

with open('Bassin_percent.json', 'w') as outfile:
    json.dump(dict_percent, outfile)


wookbook = openpyxl.load_workbook("Données/PoleEmploi/Extraction_OFF_1_EPCI.xlsx")

wookbook.active = 0

sheet  = wookbook.active

first_rows = []

for row in sheet.iter_rows(max_row=3):
    first_rows.append(row)

dict_number = {'01': {}, '38': {}, '69': {}, '73': {}, '74': {}}
dict_percent = {'01': {}, '38': {}, '69': {}, '73': {}, '74': {}}

row_data = []

for row in sheet.iter_rows(max_row=3):
    this_data = []
    for i in range(7, len(row)):       
        this_data.append(row[i].value)
    row_data.append(this_data)

row_data[2] = row_data[2][2:]



for row in sheet.iter_rows(min_row=4):
    if not dict_number.get(row[0].value).get(row[2].value):
        dict_number[row[0].value][row[2].value] = {'PE': {}, 'PE-CUMUL12MOIS': {}, 'TOFF': {}, 'TOFF-CUMUL12MOIS': {}}


index = 0
for row in sheet.iter_rows(min_row=4):
    for i in range(0, len(row_data[2]), 2):
           
        dict_number[row[0].value][row[2].value][row[6].value][row_data[2][i]] = row[i+9].value
    dict_number[row[0].value][row[2].value][row[6].value]["nbr_offres"] = row[7].value


for row in sheet.iter_rows(min_row=4):
    if not dict_percent.get(row[0].value).get(row[2].value):
        dict_percent[row[0].value][row[2].value] = {'PE': {}, 'PE-CUMUL12MOIS': {}, 'TOFF': {}, 'TOFF-CUMUL12MOIS': {}}


index = 0
for row in sheet.iter_rows(min_row=4):
    for i in range(0, len(row_data[2]), 2):
        dict_percent[row[0].value][row[2].value][row[6].value][row_data[2][i]] = row[i+10].value
    dict_percent[row[0].value][row[2].value][row[6].value]["nbr_offres"] = row[7].value

with open('EPCI_number.json', 'w') as outfile:
    json.dump(dict_number, outfile)

with open('EPCI_percent.json', 'w') as outfile:
    json.dump(dict_percent, outfile)